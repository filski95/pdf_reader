import openpyxl as xl
import os
from openpyxl.styles import Font


def create_combined_excel() -> None:
    # * comining all single files into one, to ease printing of labels
    dir = "../excel_files/output_files/"
    files = os.listdir(dir)
    wb = xl.Workbook()
    ws = wb.active
    _combine_excels(dir, ws, files)
    _format_za_cells(ws)
    wb.save("../excel_files/output_files/combined.xlsx")


def _combine_excels(dir: str, ws, files: list[str]) -> None:
    for file in files:
        if file.endswith(".xlsx"):
            file_wb = xl.load_workbook(filename=os.path.join(dir + file))
            file_ws = file_wb.active
            for value in file_ws.values:  # * generator
                ws.append(value)


def _format_za_cells(ws) -> None:
    # * styling of cells with ZAs.
    za_cell_style = Font(bold=True, color="FF0000", size=14)
    for row in ws.iter_rows(min_row=1, max_col=1):
        for cell in row:
            if str(cell.value).startswith("ZA"):
                cell.font = za_cell_style
    ws.column_dimensions["A"].width = 40
